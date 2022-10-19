
# -*- coding: utf-8 -*-
import openpyxl
from os import remove
from openpyxl.styles import Alignment, PatternFill
from data import *
from main import BOOK_NAME, DATE, MONITOR_POS
from thresholdData import *
        
def writeToExcel_eb2212(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in eb2212_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if eb2212_long_change[key] - eb2212_short_change[key] >= 0:
                ALERT_LIST.append(eb2212_long_change[key] - eb2212_short_change[key])
            else: 
                ALERT_LIST.append(eb2212_short_change[key] - eb2212_long_change[key])
        except:
            pass
    for key in eb2212_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if eb2212_long_change[key] - eb2212_short_change[key] >= 0:
                ALERT_LIST.append(eb2212_long_change[key] - eb2212_short_change[key])
            else: 
                ALERT_LIST.append(eb2212_short_change[key] - eb2212_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = eb2111_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = eb2212_long.get("&nbsp;")
        total_short =  eb2212_short.get("&nbsp;")
        total_long_change = eb2212_long_change.get("&nbsp;")
        total_short_change = eb2212_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = eb2212_long.get("")
        total_short =  eb2212_short.get("")
        total_long_change = eb2212_long_change.get("")
        total_short_change = eb2212_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in eb2212_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = eb2212_long[key]
        sheet.cell(ROW, COL+2).value = eb2212_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in eb2212_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = eb2212_short[key]
        sheet.cell(ROW, COL+2).value = eb2212_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in eb2212_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = eb2212_long[key]
        try:
            sheet.cell(ROW, COL+2).value = eb2212_short[key]
            val = eb2212_long[key] - eb2212_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = eb2212_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = eb2212_short_change[key]
            val = eb2212_long_change[key] - eb2212_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in eb2212_short:
        if key == "&nbsp;" or key == "":continue
        if eb2212_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = eb2212_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = eb2212_long[key]
            val = eb2212_long[key] - eb2212_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = eb2212_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = eb2212_long_change[key]
            val = eb2212_long_change[key] - eb2212_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in eb2212_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += eb2212_long[key]
        long_5_change = eb2212_long_change[key]
    tem_count = 0
    for key in eb2212_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += eb2212_short[key]
        short_5_change = eb2212_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = eb2212_long_change[key]-eb2212_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = eb2212_long[hold]
            sheet.cell(ROW, COL+2).value = eb2212_short[hold]
            sheet.cell(ROW, COL+3).value = eb2212_long[hold]-eb2212_short[hold]
            sheet.cell(ROW, COL+4).value = eb2212_long_change[hold]
            sheet.cell(ROW, COL+5).value = eb2212_short_change[hold]
            sheet.cell(ROW, COL+6).value = eb2212_long_change[hold] - eb2212_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_eb2212(BOOK_NAME, DATE, "eb2212")

def writeToExcel_eg2301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in eg2301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if eg2301_long_change[key] - eg2301_short_change[key] >= 0:
                ALERT_LIST.append(eg2301_long_change[key] - eg2301_short_change[key])
            else: 
                ALERT_LIST.append(eg2301_short_change[key] - eg2301_long_change[key])
        except:
            pass
    for key in eg2301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if eg2301_long_change[key] - eg2301_short_change[key] >= 0:
                ALERT_LIST.append(eg2301_long_change[key] - eg2301_short_change[key])
            else: 
                ALERT_LIST.append(eg2301_short_change[key] - eg2301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = eg2201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = eg2301_long.get("&nbsp;")
        total_short =  eg2301_short.get("&nbsp;")
        total_long_change = eg2301_long_change.get("&nbsp;")
        total_short_change = eg2301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = eg2301_long.get("")
        total_short =  eg2301_short.get("")
        total_long_change = eg2301_long_change.get("")
        total_short_change = eg2301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in eg2301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = eg2301_long[key]
        sheet.cell(ROW, COL+2).value = eg2301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in eg2301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = eg2301_short[key]
        sheet.cell(ROW, COL+2).value = eg2301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in eg2301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = eg2301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = eg2301_short[key]
            val = eg2301_long[key] - eg2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = eg2301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = eg2301_short_change[key]
            val = eg2301_long_change[key] - eg2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in eg2301_short:
        if key == "&nbsp;" or key == "":continue
        if eg2301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = eg2301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = eg2301_long[key]
            val = eg2301_long[key] - eg2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = eg2301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = eg2301_long_change[key]
            val = eg2301_long_change[key] - eg2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in eg2301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += eg2301_long[key]
        long_5_change = eg2301_long_change[key]
    tem_count = 0
    for key in eg2301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += eg2301_short[key]
        short_5_change = eg2301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = eg2301_long_change[key]-eg2301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = eg2301_long[hold]
            sheet.cell(ROW, COL+2).value = eg2301_short[hold]
            sheet.cell(ROW, COL+3).value = eg2301_long[hold]-eg2301_short[hold]
            sheet.cell(ROW, COL+4).value = eg2301_long_change[hold]
            sheet.cell(ROW, COL+5).value = eg2301_short_change[hold]
            sheet.cell(ROW, COL+6).value = eg2301_long_change[hold] - eg2301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_eg2301(BOOK_NAME, DATE, "eg2301")

def writeToExcel_pg2212(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in pg2212_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if pg2212_long_change[key] - pg2212_short_change[key] >= 0:
                ALERT_LIST.append(pg2212_long_change[key] - pg2212_short_change[key])
            else: 
                ALERT_LIST.append(pg2212_short_change[key] - pg2212_long_change[key])
        except:
            pass
    for key in pg2212_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if pg2212_long_change[key] - pg2212_short_change[key] >= 0:
                ALERT_LIST.append(pg2212_long_change[key] - pg2212_short_change[key])
            else: 
                ALERT_LIST.append(pg2212_short_change[key] - pg2212_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = pg2111_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = pg2212_long.get("&nbsp;")
        total_short =  pg2212_short.get("&nbsp;")
        total_long_change = pg2212_long_change.get("&nbsp;")
        total_short_change = pg2212_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = pg2212_long.get("")
        total_short =  pg2212_short.get("")
        total_long_change = pg2212_long_change.get("")
        total_short_change = pg2212_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in pg2212_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = pg2212_long[key]
        sheet.cell(ROW, COL+2).value = pg2212_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in pg2212_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = pg2212_short[key]
        sheet.cell(ROW, COL+2).value = pg2212_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in pg2212_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = pg2212_long[key]
        try:
            sheet.cell(ROW, COL+2).value = pg2212_short[key]
            val = pg2212_long[key] - pg2212_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = pg2212_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = pg2212_short_change[key]
            val = pg2212_long_change[key] - pg2212_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in pg2212_short:
        if key == "&nbsp;" or key == "":continue
        if pg2212_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = pg2212_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = pg2212_long[key]
            val = pg2212_long[key] - pg2212_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = pg2212_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = pg2212_long_change[key]
            val = pg2212_long_change[key] - pg2212_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in pg2212_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += pg2212_long[key]
        long_5_change = pg2212_long_change[key]
    tem_count = 0
    for key in pg2212_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += pg2212_short[key]
        short_5_change = pg2212_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = pg2212_long_change[key]-pg2212_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = pg2212_long[hold]
            sheet.cell(ROW, COL+2).value = pg2212_short[hold]
            sheet.cell(ROW, COL+3).value = pg2212_long[hold]-pg2212_short[hold]
            sheet.cell(ROW, COL+4).value = pg2212_long_change[hold]
            sheet.cell(ROW, COL+5).value = pg2212_short_change[hold]
            sheet.cell(ROW, COL+6).value = pg2212_long_change[hold] - pg2212_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_pg2212(BOOK_NAME, DATE, "pg2212")

def writeToExcel_pp2301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in pp2301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if pp2301_long_change[key] - pp2301_short_change[key] >= 0:
                ALERT_LIST.append(pp2301_long_change[key] - pp2301_short_change[key])
            else: 
                ALERT_LIST.append(pp2301_short_change[key] - pp2301_long_change[key])
        except:
            pass
    for key in pp2301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if pp2301_long_change[key] - pp2301_short_change[key] >= 0:
                ALERT_LIST.append(pp2301_long_change[key] - pp2301_short_change[key])
            else: 
                ALERT_LIST.append(pp2301_short_change[key] - pp2301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = pp2201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = pp2301_long.get("&nbsp;")
        total_short =  pp2301_short.get("&nbsp;")
        total_long_change = pp2301_long_change.get("&nbsp;")
        total_short_change = pp2301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = pp2301_long.get("")
        total_short =  pp2301_short.get("")
        total_long_change = pp2301_long_change.get("")
        total_short_change = pp2301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in pp2301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = pp2301_long[key]
        sheet.cell(ROW, COL+2).value = pp2301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in pp2301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = pp2301_short[key]
        sheet.cell(ROW, COL+2).value = pp2301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in pp2301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = pp2301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = pp2301_short[key]
            val = pp2301_long[key] - pp2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = pp2301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = pp2301_short_change[key]
            val = pp2301_long_change[key] - pp2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in pp2301_short:
        if key == "&nbsp;" or key == "":continue
        if pp2301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = pp2301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = pp2301_long[key]
            val = pp2301_long[key] - pp2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = pp2301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = pp2301_long_change[key]
            val = pp2301_long_change[key] - pp2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in pp2301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += pp2301_long[key]
        long_5_change = pp2301_long_change[key]
    tem_count = 0
    for key in pp2301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += pp2301_short[key]
        short_5_change = pp2301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = pp2301_long_change[key]-pp2301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = pp2301_long[hold]
            sheet.cell(ROW, COL+2).value = pp2301_short[hold]
            sheet.cell(ROW, COL+3).value = pp2301_long[hold]-pp2301_short[hold]
            sheet.cell(ROW, COL+4).value = pp2301_long_change[hold]
            sheet.cell(ROW, COL+5).value = pp2301_short_change[hold]
            sheet.cell(ROW, COL+6).value = pp2301_long_change[hold] - pp2301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_pp2301(BOOK_NAME, DATE, "pp2301")

def writeToExcel_l2301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in l2301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if l2301_long_change[key] - l2301_short_change[key] >= 0:
                ALERT_LIST.append(l2301_long_change[key] - l2301_short_change[key])
            else: 
                ALERT_LIST.append(l2301_short_change[key] - l2301_long_change[key])
        except:
            pass
    for key in l2301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if l2301_long_change[key] - l2301_short_change[key] >= 0:
                ALERT_LIST.append(l2301_long_change[key] - l2301_short_change[key])
            else: 
                ALERT_LIST.append(l2301_short_change[key] - l2301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = l2201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = l2301_long.get("&nbsp;")
        total_short =  l2301_short.get("&nbsp;")
        total_long_change = l2301_long_change.get("&nbsp;")
        total_short_change = l2301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = l2301_long.get("")
        total_short =  l2301_short.get("")
        total_long_change = l2301_long_change.get("")
        total_short_change = l2301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in l2301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = l2301_long[key]
        sheet.cell(ROW, COL+2).value = l2301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in l2301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = l2301_short[key]
        sheet.cell(ROW, COL+2).value = l2301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in l2301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = l2301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = l2301_short[key]
            val = l2301_long[key] - l2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = l2301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = l2301_short_change[key]
            val = l2301_long_change[key] - l2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in l2301_short:
        if key == "&nbsp;" or key == "":continue
        if l2301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = l2301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = l2301_long[key]
            val = l2301_long[key] - l2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = l2301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = l2301_long_change[key]
            val = l2301_long_change[key] - l2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in l2301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += l2301_long[key]
        long_5_change = l2301_long_change[key]
    tem_count = 0
    for key in l2301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += l2301_short[key]
        short_5_change = l2301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = l2301_long_change[key]-l2301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = l2301_long[hold]
            sheet.cell(ROW, COL+2).value = l2301_short[hold]
            sheet.cell(ROW, COL+3).value = l2301_long[hold]-l2301_short[hold]
            sheet.cell(ROW, COL+4).value = l2301_long_change[hold]
            sheet.cell(ROW, COL+5).value = l2301_short_change[hold]
            sheet.cell(ROW, COL+6).value = l2301_long_change[hold] - l2301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_l2301(BOOK_NAME, DATE, "l2301")

def writeToExcel_v2301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in v2301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if v2301_long_change[key] - v2301_short_change[key] >= 0:
                ALERT_LIST.append(v2301_long_change[key] - v2301_short_change[key])
            else: 
                ALERT_LIST.append(v2301_short_change[key] - v2301_long_change[key])
        except:
            pass
    for key in v2301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if v2301_long_change[key] - v2301_short_change[key] >= 0:
                ALERT_LIST.append(v2301_long_change[key] - v2301_short_change[key])
            else: 
                ALERT_LIST.append(v2301_short_change[key] - v2301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = v2201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = v2301_long.get("&nbsp;")
        total_short =  v2301_short.get("&nbsp;")
        total_long_change = v2301_long_change.get("&nbsp;")
        total_short_change = v2301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = v2301_long.get("")
        total_short =  v2301_short.get("")
        total_long_change = v2301_long_change.get("")
        total_short_change = v2301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in v2301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = v2301_long[key]
        sheet.cell(ROW, COL+2).value = v2301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in v2301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = v2301_short[key]
        sheet.cell(ROW, COL+2).value = v2301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in v2301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = v2301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = v2301_short[key]
            val = v2301_long[key] - v2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = v2301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = v2301_short_change[key]
            val = v2301_long_change[key] - v2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in v2301_short:
        if key == "&nbsp;" or key == "":continue
        if v2301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = v2301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = v2301_long[key]
            val = v2301_long[key] - v2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = v2301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = v2301_long_change[key]
            val = v2301_long_change[key] - v2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in v2301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += v2301_long[key]
        long_5_change = v2301_long_change[key]
    tem_count = 0
    for key in v2301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += v2301_short[key]
        short_5_change = v2301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = v2301_long_change[key]-v2301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = v2301_long[hold]
            sheet.cell(ROW, COL+2).value = v2301_short[hold]
            sheet.cell(ROW, COL+3).value = v2301_long[hold]-v2301_short[hold]
            sheet.cell(ROW, COL+4).value = v2301_long_change[hold]
            sheet.cell(ROW, COL+5).value = v2301_short_change[hold]
            sheet.cell(ROW, COL+6).value = v2301_long_change[hold] - v2301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_v2301(BOOK_NAME, DATE, "v2301")

def writeToExcel_MA301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in MA301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if MA301_long_change[key] - MA301_short_change[key] >= 0:
                ALERT_LIST.append(MA301_long_change[key] - MA301_short_change[key])
            else: 
                ALERT_LIST.append(MA301_short_change[key] - MA301_long_change[key])
        except:
            pass
    for key in MA301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if MA301_long_change[key] - MA301_short_change[key] >= 0:
                ALERT_LIST.append(MA301_long_change[key] - MA301_short_change[key])
            else: 
                ALERT_LIST.append(MA301_short_change[key] - MA301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = MA201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = MA301_long.get("&nbsp;")
        total_short =  MA301_short.get("&nbsp;")
        total_long_change = MA301_long_change.get("&nbsp;")
        total_short_change = MA301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = MA301_long.get("")
        total_short =  MA301_short.get("")
        total_long_change = MA301_long_change.get("")
        total_short_change = MA301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in MA301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = MA301_long[key]
        sheet.cell(ROW, COL+2).value = MA301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in MA301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = MA301_short[key]
        sheet.cell(ROW, COL+2).value = MA301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in MA301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = MA301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = MA301_short[key]
            val = MA301_long[key] - MA301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = MA301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = MA301_short_change[key]
            val = MA301_long_change[key] - MA301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in MA301_short:
        if key == "&nbsp;" or key == "":continue
        if MA301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = MA301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = MA301_long[key]
            val = MA301_long[key] - MA301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = MA301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = MA301_long_change[key]
            val = MA301_long_change[key] - MA301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in MA301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += MA301_long[key]
        long_5_change = MA301_long_change[key]
    tem_count = 0
    for key in MA301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += MA301_short[key]
        short_5_change = MA301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = MA301_long_change[key]-MA301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = MA301_long[hold]
            sheet.cell(ROW, COL+2).value = MA301_short[hold]
            sheet.cell(ROW, COL+3).value = MA301_long[hold]-MA301_short[hold]
            sheet.cell(ROW, COL+4).value = MA301_long_change[hold]
            sheet.cell(ROW, COL+5).value = MA301_short_change[hold]
            sheet.cell(ROW, COL+6).value = MA301_long_change[hold] - MA301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_MA301(BOOK_NAME, DATE, "MA301")

def writeToExcel_TA301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in TA301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if TA301_long_change[key] - TA301_short_change[key] >= 0:
                ALERT_LIST.append(TA301_long_change[key] - TA301_short_change[key])
            else: 
                ALERT_LIST.append(TA301_short_change[key] - TA301_long_change[key])
        except:
            pass
    for key in TA301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if TA301_long_change[key] - TA301_short_change[key] >= 0:
                ALERT_LIST.append(TA301_long_change[key] - TA301_short_change[key])
            else: 
                ALERT_LIST.append(TA301_short_change[key] - TA301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = TA201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = TA301_long.get("&nbsp;")
        total_short =  TA301_short.get("&nbsp;")
        total_long_change = TA301_long_change.get("&nbsp;")
        total_short_change = TA301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = TA301_long.get("")
        total_short =  TA301_short.get("")
        total_long_change = TA301_long_change.get("")
        total_short_change = TA301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in TA301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = TA301_long[key]
        sheet.cell(ROW, COL+2).value = TA301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in TA301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = TA301_short[key]
        sheet.cell(ROW, COL+2).value = TA301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in TA301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = TA301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = TA301_short[key]
            val = TA301_long[key] - TA301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = TA301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = TA301_short_change[key]
            val = TA301_long_change[key] - TA301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in TA301_short:
        if key == "&nbsp;" or key == "":continue
        if TA301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = TA301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = TA301_long[key]
            val = TA301_long[key] - TA301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = TA301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = TA301_long_change[key]
            val = TA301_long_change[key] - TA301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in TA301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += TA301_long[key]
        long_5_change = TA301_long_change[key]
    tem_count = 0
    for key in TA301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += TA301_short[key]
        short_5_change = TA301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = TA301_long_change[key]-TA301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = TA301_long[hold]
            sheet.cell(ROW, COL+2).value = TA301_short[hold]
            sheet.cell(ROW, COL+3).value = TA301_long[hold]-TA301_short[hold]
            sheet.cell(ROW, COL+4).value = TA301_long_change[hold]
            sheet.cell(ROW, COL+5).value = TA301_short_change[hold]
            sheet.cell(ROW, COL+6).value = TA301_long_change[hold] - TA301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_TA301(BOOK_NAME, DATE, "TA301")

def writeToExcel_PF211(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in PF211_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if PF211_long_change[key] - PF211_short_change[key] >= 0:
                ALERT_LIST.append(PF211_long_change[key] - PF211_short_change[key])
            else: 
                ALERT_LIST.append(PF211_short_change[key] - PF211_long_change[key])
        except:
            pass
    for key in PF211_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if PF211_long_change[key] - PF211_short_change[key] >= 0:
                ALERT_LIST.append(PF211_long_change[key] - PF211_short_change[key])
            else: 
                ALERT_LIST.append(PF211_short_change[key] - PF211_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = PF111_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = PF211_long.get("&nbsp;")
        total_short =  PF211_short.get("&nbsp;")
        total_long_change = PF211_long_change.get("&nbsp;")
        total_short_change = PF211_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = PF211_long.get("")
        total_short =  PF211_short.get("")
        total_long_change = PF211_long_change.get("")
        total_short_change = PF211_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in PF211_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = PF211_long[key]
        sheet.cell(ROW, COL+2).value = PF211_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in PF211_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = PF211_short[key]
        sheet.cell(ROW, COL+2).value = PF211_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in PF211_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = PF211_long[key]
        try:
            sheet.cell(ROW, COL+2).value = PF211_short[key]
            val = PF211_long[key] - PF211_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = PF211_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = PF211_short_change[key]
            val = PF211_long_change[key] - PF211_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in PF211_short:
        if key == "&nbsp;" or key == "":continue
        if PF211_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = PF211_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = PF211_long[key]
            val = PF211_long[key] - PF211_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = PF211_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = PF211_long_change[key]
            val = PF211_long_change[key] - PF211_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in PF211_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += PF211_long[key]
        long_5_change = PF211_long_change[key]
    tem_count = 0
    for key in PF211_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += PF211_short[key]
        short_5_change = PF211_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = PF211_long_change[key]-PF211_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = PF211_long[hold]
            sheet.cell(ROW, COL+2).value = PF211_short[hold]
            sheet.cell(ROW, COL+3).value = PF211_long[hold]-PF211_short[hold]
            sheet.cell(ROW, COL+4).value = PF211_long_change[hold]
            sheet.cell(ROW, COL+5).value = PF211_short_change[hold]
            sheet.cell(ROW, COL+6).value = PF211_long_change[hold] - PF211_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_PF211(BOOK_NAME, DATE, "PF211")

def writeToExcel_lu2211(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in lu2211_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if lu2211_long_change[key] - lu2211_short_change[key] >= 0:
                ALERT_LIST.append(lu2211_long_change[key] - lu2211_short_change[key])
            else: 
                ALERT_LIST.append(lu2211_short_change[key] - lu2211_long_change[key])
        except:
            pass
    for key in lu2211_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if lu2211_long_change[key] - lu2211_short_change[key] >= 0:
                ALERT_LIST.append(lu2211_long_change[key] - lu2211_short_change[key])
            else: 
                ALERT_LIST.append(lu2211_short_change[key] - lu2211_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = lu2111_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = lu2211_long.get("&nbsp;")
        total_short =  lu2211_short.get("&nbsp;")
        total_long_change = lu2211_long_change.get("&nbsp;")
        total_short_change = lu2211_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = lu2211_long.get("")
        total_short =  lu2211_short.get("")
        total_long_change = lu2211_long_change.get("")
        total_short_change = lu2211_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in lu2211_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = lu2211_long[key]
        sheet.cell(ROW, COL+2).value = lu2211_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in lu2211_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = lu2211_short[key]
        sheet.cell(ROW, COL+2).value = lu2211_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in lu2211_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = lu2211_long[key]
        try:
            sheet.cell(ROW, COL+2).value = lu2211_short[key]
            val = lu2211_long[key] - lu2211_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = lu2211_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = lu2211_short_change[key]
            val = lu2211_long_change[key] - lu2211_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in lu2211_short:
        if key == "&nbsp;" or key == "":continue
        if lu2211_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = lu2211_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = lu2211_long[key]
            val = lu2211_long[key] - lu2211_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = lu2211_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = lu2211_long_change[key]
            val = lu2211_long_change[key] - lu2211_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in lu2211_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += lu2211_long[key]
        long_5_change = lu2211_long_change[key]
    tem_count = 0
    for key in lu2211_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += lu2211_short[key]
        short_5_change = lu2211_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = lu2211_long_change[key]-lu2211_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = lu2211_long[hold]
            sheet.cell(ROW, COL+2).value = lu2211_short[hold]
            sheet.cell(ROW, COL+3).value = lu2211_long[hold]-lu2211_short[hold]
            sheet.cell(ROW, COL+4).value = lu2211_long_change[hold]
            sheet.cell(ROW, COL+5).value = lu2211_short_change[hold]
            sheet.cell(ROW, COL+6).value = lu2211_long_change[hold] - lu2211_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_lu2211(BOOK_NAME, DATE, "lu2211")

def writeToExcel_fu2301(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_LIST = []
    for key in fu2301_long_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if fu2301_long_change[key] - fu2301_short_change[key] >= 0:
                ALERT_LIST.append(fu2301_long_change[key] - fu2301_short_change[key])
            else: 
                ALERT_LIST.append(fu2301_short_change[key] - fu2301_long_change[key])
        except:
            pass
    for key in fu2301_short_change:
        if key == "" or key == "&nbsp;":
            continue
        try:
            if fu2301_long_change[key] - fu2301_short_change[key] >= 0:
                ALERT_LIST.append(fu2301_long_change[key] - fu2301_short_change[key])
            else: 
                ALERT_LIST.append(fu2301_short_change[key] - fu2301_long_change[key])
        except:
            pass
    ALERT_LIST = list(set(ALERT_LIST))
    ALERT_VALUE = sum(ALERT_LIST)/len(ALERT_LIST)
    ALERT_VALUE2 = fu2201_90
    ALERT_long = []
    ALERT_short = []
    try:
        total_long = fu2301_long.get("&nbsp;")
        total_short =  fu2301_short.get("&nbsp;")
        total_long_change = fu2301_long_change.get("&nbsp;")
        total_short_change = fu2301_short_change.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = fu2301_long.get("")
        total_short =  fu2301_short.get("")
        total_long_change = fu2301_long_change.get("")
        total_short_change = fu2301_short_change.get("")

    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var 
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    for key in fu2301_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = fu2301_long[key]
        sheet.cell(ROW, COL+2).value = fu2301_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in fu2301_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = fu2301_short[key]
        sheet.cell(ROW, COL+2).value = fu2301_short_change[key]
        ROW+=1
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头 当天阈值"
    sheet.cell(ROW, COL+7).value="多头-空头 历史阈值"
    COL = 12
    ROW = 4
    for key in fu2301_long:
        if key == "&nbsp;" or key == "":continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = fu2301_long[key]
        try:
            sheet.cell(ROW, COL+2).value = fu2301_short[key]
            val = fu2301_long[key] - fu2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = fu2301_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = fu2301_short_change[key]
            val = fu2301_long_change[key] - fu2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    for key in fu2301_short:
        if key == "&nbsp;" or key == "":continue
        if fu2301_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = fu2301_short[key]
        else:continue
        try:
            sheet.cell(ROW, COL+1).value = fu2301_long[key]
            val = fu2301_long[key] - fu2301_short[key]
            sheet.cell(ROW, COL+3).value = val
        except:
            pass
        sheet.cell(ROW, COL+4).value = fu2301_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = fu2301_long_change[key]
            val = fu2301_long_change[key] - fu2301_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            sheet.cell(ROW, COL+7).value = val
            if abs(val) > ALERT_VALUE:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
            if abs(val) > ALERT_VALUE2:
                sheet.cell(ROW, COL+7).fill = ALERT
                #ALERT_long.append(key)
        except:
            pass
        ROW+=1
    sheet.cell(ROW+1, COL+5).value = "异常阈值"
    sheet.cell(ROW+1, COL+6).value = ALERT_VALUE
    sheet.cell(ROW+1, COL+7).value = ALERT_VALUE2
    COL = 4
    ROW = 60
    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"
    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前20仓位总变量"
    sheet.cell(ROW+1, COL+1).value = total_long_change
    sheet.cell(ROW+1, COL+2).value = total_short_change
    sheet.cell(ROW+2, COL).value = "前20仓位多空倾向"
    if total_long_change >= total_short_change:sheet.cell(ROW+2, COL+1).value="做多 轧差量为"+str(total_long-total_short)
    else:sheet.cell(ROW+2, COL+1).value="做空 轧差量为"+str(total_short-total_long)
    long_5 = 0
    long_5_change = 0
    short_5 = 0
    short_5_change = 0
    tem_count = 0
    for key in fu2301_long:
        if tem_count == 5:break
        else:tem_count+=1
        long_5 += fu2301_long[key]
        long_5_change = fu2301_long_change[key]
    tem_count = 0
    for key in fu2301_short:
        if tem_count == 5:break
        else:tem_count+=1
        short_5 += fu2301_short[key]
        short_5_change = fu2301_short_change[key]
    sheet.cell(ROW+3, COL).value = "前5仓位总量"
    sheet.cell(ROW+3, COL+1).value = long_5
    sheet.cell(ROW+3, COL+2).value = short_5
    sheet.cell(ROW+4, COL).value = "前5仓位总变量"
    sheet.cell(ROW+4, COL+1).value = long_5_change
    sheet.cell(ROW+4, COL+2).value = short_5_change 
    sheet.cell(ROW+5, COL).value = "前5仓位集中度"
    sheet.cell(ROW+6, COL).value = "前5仓位多空倾向"
    sheet.cell(ROW+5, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+5, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"
    if (long_5/total_long) >= (short_5/total_short):sheet.cell(ROW+6, COL+1).value = "做多 轧差量为"+str(long_5-short_5)
    else:sheet.cell(ROW+6, COL+1).value = "做空 轧差量为"+str(short_5-long_5)
    COL = 10
    ROW = 60
    sheet.cell(ROW-1, COL).value = "异常值"
    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"
    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = fu2301_long_change[key]-fu2301_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1
    COL = 4
    ROW = 80
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"
    for hold in MONITOR_POS:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = fu2301_long[hold]
            sheet.cell(ROW, COL+2).value = fu2301_short[hold]
            sheet.cell(ROW, COL+3).value = fu2301_long[hold]-fu2301_short[hold]
            sheet.cell(ROW, COL+4).value = fu2301_long_change[hold]
            sheet.cell(ROW, COL+5).value = fu2301_short_change[hold]
            sheet.cell(ROW, COL+6).value = fu2301_long_change[hold] - fu2301_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
    workbook.save(book_name)
writeToExcel_fu2301(BOOK_NAME, DATE, "fu2301")

def dataSum(book_name, date, sheet_name):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    sheets = workbook.sheetnames 
    sheet=workbook[sheets[0]] 
    sheet.title = sheet_name
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    TITLE = PatternFill('solid', fgColor="faebd7")
    LONG = PatternFill('solid', fgColor="ffc7ce")
    SHORT = PatternFill('solid', fgColor="7fffd4")
    ROW_W = 5
    COL_W = 4
    sheet.merge_cells('E3:G3')
    sheet.cell(ROW_W-2, COL_W+1).value = "主流化工品种的多空仓增减"
    sheet.merge_cells('J3:L3')
    sheet.cell(ROW_W-2, COL_W+6).value = "主流化工品种的异常持仓增减量"
    for i in range(1, len(sheets)):
        ws = workbook[sheets[i]]
        var = ws.title
        ROW_in = ROW_W
        COL_in = COL_W

        sheet.cell(ROW_in, COL_in-1).value = var
        for col in range(3, 8):
            ROW_W = ROW_in
            for row in range(59, 67):
                sheet.cell(ROW_W, COL_W).value = ws.cell(row,col).value
                ROW_W+=1
            COL_W+=1
        

        for col in range(9, 13):
            ROW_W = ROW_in
            tem_count = 0
            for row in range(59, 79):
                if ws.cell(row,col).value == None and tem_count>6:
                    break
                else:
                    sheet.cell(ROW_W, COL_W).value = ws.cell(row,col).value
                    if sheet.cell(ROW_W, COL_W).value == "做多":
                        sheet.cell(ROW_W, COL_W).fill = LONG
                    elif sheet.cell(ROW_W, COL_W).value == "做空":
                        sheet.cell(ROW_W, COL_W).fill = SHORT
                ROW_W+=1
                tem_count+=1
            COL_W+=1
        COL_W = COL_in
        ROW_W += 3
    workbook.save(BOOK_NAME)

dataSum(BOOK_NAME, DATE, "数据汇总")
