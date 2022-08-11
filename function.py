import openpyxl
from openpyxl.styles import Alignment, PatternFill
from data import *
from main import BOOK_NAME, DATE, ALERT_HOLD


def writeToExcel(book_name, date, var):
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_long = []
    ALERT_short = []

    try:
        total_long = MA209_long.get("&nbsp;")
        total_short =  MA209_short.get("&nbsp;")
        if total_long == None:raise TypeError
    except:
        total_long = MA209_long.get("")
        total_short =  MA209_short.get("")


    #排名变化
    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var # var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"

    #多头
    ROW=4
    for key in MA209_long:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = MA209_long[key]
        sheet.cell(ROW, COL+2).value = MA209_long_change[key]
        ROW+=1
    COL=7    
    ROW=4
    for key in MA209_short:
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key
        sheet.cell(ROW, COL+1).value = MA209_short[key]
        sheet.cell(ROW, COL+2).value = MA209_short_change[key]
        ROW+=1

    # 同一公司变化
    COL = 12
    ROW = 3
    sheet.cell(ROW, COL-1).value=var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="空头持仓"
    sheet.cell(ROW, COL+3).value="多头-空头"
    sheet.cell(ROW, COL+4).value="多头变量"
    sheet.cell(ROW, COL+5).value="空头变量"
    sheet.cell(ROW, COL+6).value="多头-空头"


    #多头方向看
    COL = 12
    ROW = 4
    for key in MA209_long:
        if key == "&nbsp;" or key == "":
            continue
        sheet.cell(ROW, COL).value = key
        sheet.cell(ROW, COL+1).value = MA209_long[key]

        try:
            sheet.cell(ROW, COL+2).value = MA209_short[key]
            val = MA209_long[key] - MA209_short[key]
            sheet.cell(ROW, COL+3).value = val
            # if abs(val) > 5000:
            #     sheet.cell(ROW, COL+3).fill = ALERT
        except:
            pass

        sheet.cell(ROW, COL+4).value = MA209_long_change[key]
        try:
            sheet.cell(ROW, COL+5).value = MA209_short_change[key]
            val = MA209_long_change[key] - MA209_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            if abs(val) > 5000:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_long.append(key)
        except:
            pass
        ROW+=1

    #空头方向看
    for key in MA209_short:
        if key == "&nbsp;" or key == "":
            continue
        if MA209_long.get(key) == None:
            sheet.cell(ROW, COL).value = key
            sheet.cell(ROW, COL+2).value = MA209_short[key]
        else:
            continue

        try:
            sheet.cell(ROW, COL+1).value = MA209_long[key]
            val = MA209_long[key] - MA209_short[key]
            sheet.cell(ROW, COL+3).value = val
            # if abs(val) > 5000:
            #     sheet.cell(ROW, COL+3).fill = ALERT
        except:
            pass

        sheet.cell(ROW, COL+4).value = MA209_short_change[key]
        try:
            sheet.cell(ROW, COL+5).value = MA209_long_change[key]
            val = MA209_long_change[key] - MA209_short_change[key]
            sheet.cell(ROW, COL+6).value = val
            if abs(val) > 5000:
                sheet.cell(ROW, COL+6).fill = ALERT
                ALERT_short.append(key)
        except:
            pass
        ROW+=1





    #数据小结

    COL = 4
    ROW = 35

    sheet.cell(ROW-1, COL+1).value = "多仓"
    sheet.cell(ROW-1, COL+2).value = "空仓"

    sheet.cell(ROW, COL).value = "前20仓位总量"
    sheet.cell(ROW, COL+1).value = total_long
    sheet.cell(ROW, COL+2).value = total_short
    sheet.cell(ROW+1, COL).value = "前5仓位集中度"
    sheet.cell(ROW+2, COL).value = "前5仓位多空倾向"

    long_5 = 0
    short_5 = 0
    tem_count = 0
    for key in MA209_long:
        if tem_count == 5:
            break
        else:
            tem_count+=1
        long_5 += MA209_long[key]
    tem_count = 0
    for key in MA209_short:
        if tem_count == 5:
            break
        else:
            tem_count+=1
        short_5 += MA209_short[key]
    sheet.cell(ROW+1, COL+1).value = str(round(long_5/total_long, 3) * 100)+"%"
    sheet.cell(ROW+1, COL+2).value = str(round(short_5/total_short, 3) * 100)+"%"

    sheet.merge_cells(start_row=ROW+2, start_column=COL+1, end_row=ROW+2, end_column=COL+2)
    if (long_5/total_long) >= (short_5/total_short):
        sheet.cell(ROW+2, COL+2).value = "做多 轧差量为"+str(long_5-short_5)
    else:
        sheet.cell(ROW+2, COL+1).value = "做空 轧差量为"+str(short_5-long_5)

    
    COL = 10
    ROW = 35

    sheet.cell(ROW-1, COL+1).value = "方向"
    sheet.cell(ROW-1, COL+2).value = "数量"

    for key in ALERT_long+ALERT_short:
        sheet.cell(ROW, COL).value = key
        diff = MA209_long_change[key]-MA209_short_change[key]
        if diff >= 0:
            sheet.cell(ROW, COL+1).value = "做多"
            sheet.cell(ROW, COL+2).value = diff
        else:
            sheet.cell(ROW, COL+1).value = "做空"
            sheet.cell(ROW, COL+2).value = diff
        ROW+=1


    COL = 4
    ROW = 45
    # sheet.merge_cells(start_row=ROW-1, start_column=COL+1, end_row=ROW+1, end_column=COL+10)
    sheet.cell(ROW-1, COL).value = "重点席位信息"
    sheet.cell(ROW-1, COL+1).value="多头持仓"
    sheet.cell(ROW-1, COL+2).value="空头持仓"
    sheet.cell(ROW-1, COL+3).value="多头-空头"
    sheet.cell(ROW-1, COL+4).value="多头变量"
    sheet.cell(ROW-1, COL+5).value="空头变量"
    sheet.cell(ROW-1, COL+6).value="多头-空头"

    for hold in ALERT_HOLD:
        sheet.cell(ROW, COL).value = hold
        try:
            sheet.cell(ROW, COL+1).value = MA209_long[hold]
            sheet.cell(ROW, COL+2).value = MA209_short[hold]
            sheet.cell(ROW, COL+3).value = MA209_long[hold]-MA209_short[hold]
            sheet.cell(ROW, COL+4).value = MA209_long_change[hold]
            sheet.cell(ROW, COL+5).value = MA209_short_change[hold]
            sheet.cell(ROW, COL+6).value = MA209_long_change[hold] - MA209_short_change[hold]
        except:
            sheet.cell(ROW, COL+1).value = "暂无数据"
        ROW +=1
            


    workbook.save(book_name)




writeToExcel(BOOK_NAME, DATE, "MA209")
