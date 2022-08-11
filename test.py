



from main import VAR_LIST, ALERT_HOLD


def publicHeadGenerate():
    with open("result.py",'a') as funciontFile:
        funciontFile.write("""
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from data import *
from main import BOOK_NAME, DATE, ALERT_HOLD
""")


def functionGenetate(var):
    with open("result.py",'a') as funciontFile:
        funciontFile.write("def writeToExcel_"+var+"(book_name, date, var):")
        funciontFile.write("""
    try:workbook = openpyxl.load_workbook(book_name)
    except: workbook = openpyxl.Workbook()
    try:sheet = workbook.create_sheet(var)
    except:sheet = workbook.active
    sheet.title = var
    sheet.merge_cells('A1:R1')
    sheet.cell(1,1).value = '化工数据'+date+"汇总"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ALERT = PatternFill('solid', fgColor="ffc7ce")
    ALERT_long = []
    ALERT_short = []
    try:
""")
        funciontFile.write("        total_long = "+var+"_long.get(\"&nbsp;\")\n")
        funciontFile.write("        total_short = "+var+"_short.get(\"&nbsp;\")")
        funciontFile.write("""        
        if total_long == None:raise TypeError
    except:\n""")
        funciontFile.write("        total_long = "+var+"_long.get(\"&nbsp;\")\n")
        funciontFile.write("        total_short = "+var+"_short.get(\"&nbsp;\")")
        funciontFile.write("""
    COL = 4
    ROW = 3
    sheet.cell(ROW, COL-1).value=var # var
    sheet.cell(ROW, COL).value="期货公司"
    sheet.cell(ROW, COL+1).value="多头持仓"
    sheet.cell(ROW, COL+2).value="多头变量"
    sheet.cell(ROW, COL+3).value="期货公司"
    sheet.cell(ROW, COL+4).value="空头持仓"
    sheet.cell(ROW, COL+5).value="空头变量"
    ROW=4
    """)
        funciontFile.write("        for key in "+var+"_long:")
        funciontFile.write("""
        if key == "&nbsp;" or key == "":tem_key = "总计"
        else:tem_key=key
        sheet.cell(ROW, COL).value = tem_key""")
        funciontFile.write("        sheet.cell(ROW, COL+1).value = "+var+"_long[key]")
        funciontFile.write("        sheet.cell(ROW, COL+2).value = "+var+"_long_change[key]")
        funciontFile.write("""
        ROW+=1
    COL=7    
    ROW=4""")
        # funciontFile.write()




def generateMain(VAR_LIST):
    publicHeadGenerate()
    # for var in VAR_LIST:
    functionGenetate("MA209")


generateMain(VAR_LIST)